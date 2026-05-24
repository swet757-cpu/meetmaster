from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
SOURCE_XLSX = BASE_DIR / "dds_dashboard_2026-05-13_v2.xlsx"
TEMPLATE_HTML = BASE_DIR / "dashboard.html"
OUTPUT_HTML = BASE_DIR / "dds_dashboard_2026-05-13_v2.html"
INITIAL_BALANCE = 1_296_288.48


def normalize_date(value) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def load_operations() -> list[dict[str, object]]:
    wb = load_workbook(SOURCE_XLSX, data_only=False)
    ws = wb["Ops"]
    operations: list[dict[str, object]] = []

    for row_idx in range(4, ws.max_row + 1):
        row_date = normalize_date(ws.cell(row_idx, 1).value)
        operation = ws.cell(row_idx, 2).value
        article = ws.cell(row_idx, 3).value
        inflow = ws.cell(row_idx, 4).value
        outflow = ws.cell(row_idx, 5).value

        if not row_date:
            continue
        if not any(value not in (None, "", 0) for value in [inflow, outflow]):
            continue

        inflow_value = float(inflow or 0)
        outflow_value = float(outflow or 0)
        operations.append(
            {
                "date": row_date,
                "operation": str(operation or "").strip() or ("Поступление" if inflow_value > 0 else "Списание"),
                "article": str(article or "").strip() or "Без статьи",
                "inflow": inflow_value,
                "outflow": outflow_value,
                "net": inflow_value - outflow_value,
            }
        )

    operations.sort(key=lambda item: (item["date"], item["operation"], item["article"]))
    return operations


def main() -> None:
    html = TEMPLATE_HTML.read_text(encoding="utf-8")
    report = {"initialBalance": INITIAL_BALANCE, "operations": load_operations()}
    replacement = "const report = " + json.dumps(report, ensure_ascii=False, separators=(",", ":")) + ";"
    html = re.sub(r"const report = .*?;\nconst baseOps =", replacement + "\nconst baseOps =", html, count=1, flags=re.S)
    html = html.replace("a.download = 'dds_added_operations.csv';", "a.download = 'dds_added_operations_2026.csv';")
    OUTPUT_HTML.write_text(html, encoding="utf-8")
    print(f"Создан HTML: {OUTPUT_HTML.name}")
    print(f"Операций выгружено: {len(report['operations'])}")


if __name__ == "__main__":
    main()
